import pandas as pd
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from time import sleep
import undetected_chromedriver as uc
from openpyxl import Workbook
from datetime import datetime
import requests
from urllib.parse import quote
from openpyxl import load_workbook
import pytest

prompts = "give me 10 weird inputs having same intents for the following example statement: Input: {} Intent: {}"
intent = "decr_ac_temp"

op = webdriver.ChromeOptions()
# op.add_argument(f"user-agent={UserAgent.random}")
op.add_argument("user-data-dir=./")
op.add_experimental_option("detach", True)
op.add_experimental_option("excludeSwitches", ["enable-logging"])

# driver = uc.Chrome(chrome_options=op)
#
# MAIL = "*********************"
# PASSWORD = "***********"
# PATH = "chromedriver"
#
# driver.get('https://chat.openai.com')


# inputElements = driver.find_elements(By.TAG_NAME, "button")
# inputElements[0].click()
#
# sleep(3)
#
# mail = driver.find_elements(By.TAG_NAME,"input")[1]
# mail.send_keys(MAIL)
#
# btn=driver.find_elements(By.TAG_NAME,"button")[0]
# btn.click()
#
# password= driver.find_elements(By.TAG_NAME,"input")[2]
# password.send_keys(PASSWORD)

# sleep(3)

# wait = WebDriverWait(driver, 10)
# btn = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "_button-login-password")))
# btn.click()
# sleep(10)


question_response = {}

results = []
responses = []
headers = {'accept': 'application/json'}

workbook_list = load_workbook('Prompts_Intention_List.xlsx')
sheet_list = workbook_list.active

for row in sheet_list.iter_rows(min_row=1, values_only=True):
    driver = uc.Chrome(chrome_options=op)

    MAIL = "*********************"
    PASSWORD = "***********"
    PATH = "chromedriver"

    driver.get('https://chat.openai.com')

    inputElements = driver.find_elements(By.TAG_NAME, "textarea")
    intent, prompt, intention = row[:3]  # Assuming the data is in the first three columns

    inputElements[0].send_keys(prompts.format(prompt, intention))
    sleep(2)
    inputElements[0].send_keys(Keys.ENTER)
    sleep(20)
    inputElements = driver.find_elements(By.TAG_NAME, "li")
    sleep(2)
    # for element in inputElements:
    #     results.append(element.text)
    #     encoded_string = quote(element.text)
    #     url = 'http://192.168.1.221:8088/get_intent/?text=' + encoded_string
    #     response = requests.post(url, headers=headers)
    #     if intent not in response.text:
    #         print("Incorrect response for question: " + element.text)
    #     else:
    #         print("Found correct response")
    #     responses.append(response)
    #     question_response[element.text] = response

    driver.quit()

        # wb = Workbook()
        #
        # # Select the active worksheet
        # ws = wb.active
        #
        # df = pd.DataFrame(question_response)
        #
        # timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        #
        # header = list(question_response.keys())
        # ws.append(header)
        #
        # # Write the data rows
        # for row_data in zip(*question_response.values()):
        #     ws.append(row_data)
        #
        # filename = f"output_responses/output_{timestamp}_{intent}.xlsx"
        #
        # df.to_excel(filename, index=False)
        #
        # print("Excel file saved successfully.")
        # sleep(3)
        # df = pd.read_excel(filename)
        #
        # # Transpose the DataFrame to convert rows to columns
        # df_transposed = df.transpose()
        # sleep(2)
        # # Save the transposed DataFrame to a new Excel file
        # df_transposed.to_excel(filename, header=False)

print("Excel file saved successfully with rows converted to columns.")
print(results)
